"""
Сервис обработки и хранения файлов
"""
import mimetypes
import os
import tempfile
import base64
import csv
import shutil
import uuid
from typing import List, Dict, Any, Optional, Tuple
from fastapi import UploadFile
import pdfplumber
from ..core.config import settings
from ..core.exceptions import ValidationError


class FileService:
    """Сервис для обработки различных типов файлов"""
    
    def __init__(self):
        os.makedirs(settings.uploads_dir, exist_ok=True)
    
    def parse_pdf(self, file_path: str) -> str:
        """Извлекает текст из PDF файла"""
        text = ""
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text
        return text
    
    def parse_excel_xlsx(self, file_path: str, max_chars: int = 8000) -> str:
        """Парсит Excel файлы .xlsx"""
        try:
            from openpyxl import load_workbook
        except Exception:
            return "Не удалось загрузить модуль openpyxl для чтения Excel. Установите зависимость."
        
        try:
            wb = load_workbook(filename=file_path, data_only=True, read_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                parts.append(f"Лист: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    row_values = ["" if v is None else str(v) for v in row]
                    parts.append("\t".join(row_values))
                    if sum(len(p) for p in parts) > max_chars:
                        break
                if sum(len(p) for p in parts) > max_chars:
                    break
            text = "\n".join(parts)
            try:
                wb.close()
            except Exception:
                pass
            if len(text) > max_chars:
                text = text[:max_chars] + "\n... (обрезано)"
            return text
        except Exception as e:
            return f"Ошибка чтения Excel: {e}"
    
    def parse_csv(self, file_path: str, max_chars: int = 8000, delimiter: str = ',') -> str:
        """Парсит CSV файлы"""
        try:
            parts = []
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    parts.append(",".join([str(c) for c in row]))
                    if sum(len(p) for p in parts) > max_chars:
                        break
            text = "\n".join(parts)
            if len(text) > max_chars:
                text = text[:max_chars] + "\n... (обрезано)"
            return text
        except Exception as e:
            return f"Ошибка чтения CSV: {e}"
    
    def read_text_file(self, file_path: str, max_chars: int = 8000) -> str:
        """Читает текстовые файлы"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read(max_chars + 1)
                if len(text) > max_chars:
                    text = text[:max_chars] + "\n... (обрезано)"
                return text
        except Exception as e:
            return f"Ошибка чтения текста: {e}"
    
    def convert_pdf_to_images(self, file_path: str, max_pages: Optional[int] = None, zoom: float = 2.0) -> List[str]:
        """Конвертирует PDF в изображения base64"""
        images_base64 = []
        try:
            import fitz  # PyMuPDF
        except Exception:
            return images_base64
        
        try:
            doc = fitz.open(file_path)
            total_pages = len(doc)
            pages_to_process = total_pages if max_pages is None else min(total_pages, max_pages)
            for page_index in range(pages_to_process):
                page = doc.load_page(page_index)
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                png_bytes = pix.tobytes("png")
                images_base64.append(base64.b64encode(png_bytes).decode("utf-8"))
            doc.close()
        except Exception:
            pass
        return images_base64
    
    def parse_excel_any(self, file_path: str, suffix: str, max_chars: int = 8000) -> str:
        """Универсальный парсер Excel файлов"""
        if suffix in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            return self.parse_excel_xlsx(file_path, max_chars=max_chars)
        elif suffix == '.xls':
            return self._parse_excel_xls(file_path, max_chars=max_chars)
        elif suffix == '.xlsb':
            return self._parse_excel_xlsb(file_path, max_chars=max_chars)
        elif suffix == '.ods':
            return self._parse_ods(file_path, max_chars=max_chars)
        elif suffix == '.csv':
            return self.parse_csv(file_path, max_chars=max_chars, delimiter=',')
        elif suffix == '.tsv':
            return self.parse_csv(file_path, max_chars=max_chars, delimiter='\t')
        elif suffix == '.txt':
            return self.read_text_file(file_path, max_chars=max_chars)
        else:
            return f"Неподдерживаемый формат файла: {suffix}"
    
    def _parse_excel_xls(self, file_path: str, max_chars: int = 8000) -> str:
        """Парсит старые Excel файлы .xls"""
        try:
            import xlrd
            parts = []
            wb = xlrd.open_workbook(file_path)
            for sheet in wb.sheets():
                parts.append(f"Лист: {sheet.name}")
                for r in range(sheet.nrows):
                    row_values = ["" if sheet.cell_value(r, c) is None else str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                    parts.append("\t".join(row_values))
                    if sum(len(p) for p in parts) > max_chars:
                        break
                if sum(len(p) for p in parts) > max_chars:
                    break
            text = "\n".join(parts)
            if len(text) > max_chars:
                text = text[:max_chars] + "\n... (обрезано)"
            return text
        except Exception:
            return self._parse_with_pandas(file_path, max_chars)
    
    def _parse_excel_xlsb(self, file_path: str, max_chars: int = 8000) -> str:
        """Парсит .xlsb файлы"""
        try:
            import pandas as pd
            df = pd.read_excel(file_path, engine='pyxlsb')
            text = df.to_csv(index=False)
            if len(text) > max_chars:
                text = text[:max_chars] + "\n... (обрезано)"
            return text
        except Exception as e:
            return f"Не удалось прочитать .xlsb: {e}"
    
    def _parse_ods(self, file_path: str, max_chars: int = 8000) -> str:
        """Парсит .ods файлы"""
        try:
            import pandas as pd
            xls = pd.ExcelFile(file_path)
            parts = []
            for sheet_name in xls.sheet_names:
                parts.append(f"Лист: {sheet_name}")
                df = xls.parse(sheet_name)
                parts.append(df.to_csv(index=False))
                if sum(len(p) for p in parts) > max_chars:
                    break
            text = "\n".join(parts)
            try:
                xls.close()
            except Exception:
                pass
            if len(text) > max_chars:
                text = text[:max_chars] + "\n... (обрезано)"
            return text
        except Exception as e:
            return f"Не удалось прочитать .ods: {e}"
    
    def _parse_with_pandas(self, file_path: str, max_chars: int = 8000) -> str:
        """Парсит через pandas как универсальный фолбэк"""
        try:
            import pandas as pd
            xls = pd.ExcelFile(file_path)
            parts = []
            for sheet_name in xls.sheet_names:
                parts.append(f"Лист: {sheet_name}")
                df = xls.parse(sheet_name)
                parts.append(df.to_csv(index=False))
                if sum(len(p) for p in parts) > max_chars:
                    break
            text = "\n".join(parts)
            try:
                xls.close()
            except Exception:
                pass
            if len(text) > max_chars:
                text = text[:max_chars] + "\n... (обрезано)"
            return text
        except Exception:
            return "Не удалось прочитать файл. Установите необходимые зависимости."
    
    async def process_uploaded_files(self, files: List[Any]) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Обрабатывает загруженные файлы (legacy AI сценарии)"""
        all_documents = []
        temp_files = []
        
        for i, file in enumerate(files):
            if i >= settings.max_files:
                break
            
            content = file.read()
            if len(content) > settings.max_file_size:
                continue
            
            suffix = os.path.splitext(file.filename)[1].lower()
            # Нормализация опечаток в расширениях Excel
            excel_aliases = {
                '.xlxs': '.xlsx',
                '.xslx': '.xlsx',
                '.xlsxx': '.xlsx',
                '.exel': '.xlsx'
            }
            if suffix in excel_aliases:
                suffix = excel_aliases[suffix]
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                temp_files.append(tmp.name)
                
                if suffix in ['.jpg', '.jpeg', '.png']:
                    # Обрабатываем изображения как base64
                    b64 = base64.b64encode(content).decode('utf-8')
                    all_documents.append({
                        "filename": file.filename,
                        "type": "image",
                        "content_base64": b64
                    })
                elif suffix == '.pdf':
                    # Сохраняем PDF как файл для отправки в OpenAI
                    pdf_base64 = base64.b64encode(content).decode('utf-8')
                    all_documents.append({
                        "filename": file.filename,
                        "type": "pdf_images",  # Оставляем тип для совместимости с API
                        "content_base64": pdf_base64  # Добавляем оригинальный PDF контент
                    })
                else:
                    # Excel и другие текстовые форматы
                    if suffix in ['.xlsx', '.xlsm', '.xltx', '.xltm', '.xls', '.xlsb', '.ods', '.csv', '.tsv', '.txt']:
                        excel_text = self.parse_excel_any(tmp.name, suffix, max_chars=settings.max_prompt_length)
                        all_documents.append({
                            "filename": file.filename,
                            "type": "excel_text",
                            "text": excel_text
                        })
                    else:
                        b64 = base64.b64encode(content).decode('utf-8')
                        all_documents.append({
                            "filename": file.filename,
                            "type": "other",
                            "content_base64": b64
                        })
        
        return all_documents, temp_files
    
    def cleanup_temp_files(self, temp_files: List[str]):
        """Очищает временные файлы"""
        for f in temp_files:
            try:
                os.remove(f)
            except Exception:
                pass
    
    # === Новая функциональность хранения файлов ===
    def _safe_subdir(self, subdir: str) -> str:
        parts = [part for part in subdir.replace("\\", "/").split("/") if part and part not in (".", "..")]
        return "/".join(parts)
    
    def _normalize_filename(self, filename: str) -> str:
        base = os.path.basename(filename or "file")
        return base.replace(" ", "_")
    
    def save_upload_file(self, file: UploadFile, subdir: str) -> Dict[str, Any]:
        """
        Сохраняет загруженный файл в директорию uploads и возвращает метаданные.
        """
        safe_subdir = self._safe_subdir(subdir) or "common"
        target_dir = os.path.join(settings.uploads_dir, safe_subdir)
        os.makedirs(target_dir, exist_ok=True)
        
        original_name = self._normalize_filename(file.filename or "file")
        extension = os.path.splitext(original_name)[1]
        unique_name = f"{uuid.uuid4().hex}{extension}"
        stored_path = os.path.join(target_dir, unique_name)
        
        file.file.seek(0)
        with open(stored_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        relative_url = f"/uploads/{safe_subdir}/{unique_name}".replace("//", "/")
        size = os.path.getsize(stored_path)
        mime_type = file.content_type or mimetypes.guess_type(original_name)[0] or "application/octet-stream"
        
        return {
            "file_name": original_name,
            "stored_name": unique_name,
            "file_path": relative_url,
            "file_size": size,
            "mime_type": mime_type,
        }
    
    def delete_file(self, file_path: Optional[str]) -> None:
        """Удаляет файл из файловой системы по относительному пути /uploads/..."""
        if not file_path:
            return
        clean_path = file_path.lstrip("/\\")
        absolute_path = os.path.join(settings.project_root, clean_path)
        if os.path.commonpath([absolute_path, settings.project_root]) != settings.project_root:
            # Предотвращаем выход за пределы корня проекта
            return
        if os.path.exists(absolute_path):
            try:
                os.remove(absolute_path)
            except OSError:
                pass



